"""HTTP API for user custom sections (template, import, custom CRUD, export)."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from django.db import transaction
from django.http import FileResponse
from rest_framework import status
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import ScopedRateThrottle
from rest_framework.views import APIView

from apps.sections.export_scopes import EXPORT_SCOPE_REGISTRY, get_allowed_export_scopes
from apps.sections.models import TABLE_TO_USER_MODEL
from apps.sections.serializers import get_user_section_serializer
from apps.sections.validation import (
    assert_allowed_catalog_table,
    assert_allowed_table,
    can_insert_custom_section,
    get_catalog_model_for_table,
    get_db_header,
    headers_match_table,
    import_db_validation,
    row_dict_to_create_kwargs,
)

SECTION_IMPORT_MAX_ROWS = 5000
SECTION_EXPORT_MAX_ROWS = 10000


def _normalize_designation(value: Any) -> str:
    """Avoid Excel float artifacts (e.g. 1.0) in CharField `Designation`."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _safe_table(
    request, *, source: str = "query", custom_only: bool = True
) -> Tuple[Optional[str], Optional[Response]]:
    """Resolve `table` from query string or POST form; return (table, error_response)."""
    if source == "query":
        table = request.query_params.get("table")
    else:
        table = request.data.get("table") if hasattr(request, "data") else None
    if not table:
        return None, Response(
            {"detail": "Parameter `table` is required."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        if custom_only:
            assert_allowed_table(table)
        else:
            assert_allowed_catalog_table(table)
    except ValueError as exc:
        return None, Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    return table, None


class SectionTemplateView(APIView):
    """GET empty xlsx: sheet name = table, row 1 = expected headers (guest allowed)."""

    permission_classes = [AllowAny]

    def get(self, request):
        from openpyxl import Workbook

        table, err = _safe_table(request, source="query")
        if err:
            return err

        wb = Workbook()
        ws = wb.active
        ws.title = table
        for col, name in enumerate(get_db_header(table), start=1):
            ws.cell(row=1, column=col, value=name)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"{table}_SectionTemplate.xlsx"
        return FileResponse(
            buf,
            as_attachment=True,
            filename=filename,
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )


class SectionCatalogExportView(APIView):
    """GET xlsx export of full global catalog table (guest allowed)."""

    permission_classes = [AllowAny]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "sections_export"

    def get(self, request):
        from openpyxl import Workbook

        table, err = _safe_table(request, source="query", custom_only=False)
        if err:
            return err

        Model = get_catalog_model_for_table(table)
        headers = get_db_header(table)
        rows = list(Model.objects.values_list(*headers))

        if len(rows) > SECTION_EXPORT_MAX_ROWS:
            return Response(
                {
                    "detail": (
                        f"Too many rows to export ({len(rows)}); "
                        f"maximum is {SECTION_EXPORT_MAX_ROWS}."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        wb = Workbook()
        ws = wb.active
        ws.title = table

        for col_idx, name in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=name)

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"{table}_Catalog.xlsx"
        return FileResponse(
            buf,
            as_attachment=True,
            filename=filename,
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )


class SectionImportView(APIView):
    """POST multipart: `file` (xlsx) + `table` — bulk create user custom sections."""

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "sections_import"

    def post(self, request):
        import logging
        from openpyxl import load_workbook

        logger = logging.getLogger("sections.import")
        user = getattr(request.user, "email", None) or getattr(request.user, "username", None) or str(request.user)
        logger.info("[SectionImport] POST from user=%s | POST keys=%s | FILES keys=%s",
                    user, list(request.POST.keys()), list(request.FILES.keys()))

        table = request.POST.get("table") or request.data.get("table")
        logger.info("[SectionImport] resolved table=%r", table)
        if not table:
            logger.warning("[SectionImport] 400 — table field missing")
            return Response(
                {"detail": "Form field `table` is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            assert_allowed_table(table)
        except ValueError as exc:
            logger.warning("[SectionImport] 400 — table not allowed: %s", exc)
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        upload = request.FILES.get("file")
        logger.info("[SectionImport] upload file=%r (size=%s)",
                    getattr(upload, "name", None), getattr(upload, "size", None))
        if not upload:
            logger.warning("[SectionImport] 400 — file field missing")
            return Response(
                {"detail": "Form file field `file` is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = load_workbook(upload, data_only=True)
        except Exception as exc:  # noqa: BLE001
            logger.warning("[SectionImport] 400 — could not read workbook: %s", exc)
            return Response(
                {"detail": f"Could not read Excel file: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            logger.info("[SectionImport] workbook sheet names=%s", wb.sheetnames)
            if table not in wb.sheetnames:
                logger.warning("[SectionImport] 400 — sheet %r not found in %s", table, wb.sheetnames)
                return Response(
                    {
                        "detail": (
                            f"Worksheet named {table!r} not found in workbook "
                            f"(available sheets: {wb.sheetnames}). "
                            "The sheet tab must be named Columns, Beams, Angles, or Channels."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            sheet = wb[table]
            max_r = sheet.max_row or 1
            data_rows = max_r - 1 if max_r else 0
            if data_rows > SECTION_IMPORT_MAX_ROWS:
                return Response(
                    {
                        "detail": (
                            f"Too many data rows ({data_rows}); "
                            f"maximum is {SECTION_IMPORT_MAX_ROWS}."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            first_row = next(
                sheet.iter_rows(min_row=1, max_row=1, values_only=True),
                None,
            )
            header_row = (
                ["" if v is None else str(v) for v in first_row]
                if first_row
                else []
            )
            logger.info("[SectionImport] header_row=%s | expected=%s", header_row, get_db_header(table))
            if not headers_match_table(header_row, table):
                logger.warning("[SectionImport] 400 — header mismatch for table=%r | got=%s | expected=%s",
                               table, header_row, get_db_header(table))
                return Response(
                    {
                        "detail": "Header row does not match expected headers for this table.",
                        "expected": get_db_header(table),
                        "got": header_row,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Build a name → column-index map from the actual header row.
            # This tolerates extra columns (e.g. Id from the catalog export) at any position.
            col_map = {
                str(h).strip(): i
                for i, h in enumerate(header_row)
                if h is not None and str(h).strip()
            }

            expected = get_db_header(table)
            Ser = get_user_section_serializer(table)
            inserted = 0
            inserted_designations: List[str] = []
            ignored: List[str] = []
            rejected: List[Any] = []

            for row_idx, row in enumerate(
                sheet.iter_rows(
                    min_row=2,
                    max_row=max_r,
                    values_only=True,
                ),
                start=2,
            ):
                values = list(row)
                if all(
                    v is None or (isinstance(v, str) and str(v).strip() == "")
                    for v in values
                ):
                    continue

                raw = {
                    key: (values[col_map[key]] if col_map.get(key) is not None and col_map[key] < len(values) else None)
                    for key in expected
                }
                des_str = _normalize_designation(raw.get("Designation"))
                if not des_str:
                    rejected.append(
                        {"row": row_idx, "reason": "missing_or_empty_designation"}
                    )
                    continue
                raw["Designation"] = des_str

                row_ok = True
                for key, val in raw.items():
                    if not import_db_validation(table, key, val):
                        rejected.append(
                            {
                                "row": row_idx,
                                "designation": des_str,
                                "reason": "cell_validation_failed",
                                "key": key,
                                "value": val,
                            }
                        )
                        row_ok = False
                        break
                if not row_ok:
                    continue

                ok, code = can_insert_custom_section(table, des_str, request.user)
                if not ok:
                    if code in ("catalog_duplicate", "user_duplicate"):
                        ignored.append(des_str)
                    else:
                        rejected.append(
                            {
                                "row": row_idx,
                                "designation": des_str,
                                "reason": code,
                            }
                        )
                    continue

                payload = row_dict_to_create_kwargs(table, raw)
                ser = Ser(data=payload, context={"request": request})
                try:
                    ser.is_valid(raise_exception=True)
                    ser.save()
                    inserted += 1
                    inserted_designations.append(des_str)
                except ValidationError as exc:
                    rejected.append(
                        {
                            "row": row_idx,
                            "designation": des_str,
                            "reason": "serializer_validation",
                            "errors": exc.detail,
                        }
                    )

            logger.info("[SectionImport] done — inserted=%d ignored=%d rejected=%d",
                        inserted, len(ignored), len(rejected))
            return Response(
                {
                    "inserted": inserted,
                    "inserted_designations": inserted_designations,
                    "ignored": ignored,
                    "rejected": rejected,
                },
                status=status.HTTP_200_OK,
            )
        finally:
            wb.close()


class SectionCustomView(APIView):
    """POST JSON create one row; DELETE by designation."""

    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, FormParser]

    def post(self, request):
        table, err = _safe_table(request, source="query")
        if err:
            return err
        Ser = get_user_section_serializer(table)
        ser = Ser(data=request.data, context={"request": request})
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        instance = ser.save()
        out = Ser(instance, context={"request": request})
        return Response(out.data, status=status.HTTP_201_CREATED)

    def delete(self, request):
        table, err = _safe_table(request, source="query")
        if err:
            return err
        designation = request.query_params.get("designation")
        if not designation:
            return Response(
                {"detail": "Query parameter `designation` is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        Model = TABLE_TO_USER_MODEL[table]
        qs = Model.objects.filter(user=request.user, Designation=designation)
        total_deleted, _ = qs.delete()
        if total_deleted == 0:
            return Response(status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_204_NO_CONTENT)


class SectionCustomBulkDeleteView(APIView):
    """DELETE all user custom sections across Columns/Beams/Angles/Channels."""

    permission_classes = [IsAuthenticated]

    def delete(self, request):
        deleted: Dict[str, int] = {}
        total_deleted = 0

        with transaction.atomic():
            for table, Model in TABLE_TO_USER_MODEL.items():
                qs = Model.objects.filter(user=request.user)
                count, _ = qs.delete()
                deleted[table] = int(count)
                total_deleted += int(count)

        return Response(
            {"success": True, "deleted": deleted, "total_deleted": total_deleted},
            status=status.HTTP_200_OK,
        )


class SectionExportView(APIView):
    """GET xlsx export; v1 supports `scope=user` only (see export_scopes registry)."""

    permission_classes = [IsAuthenticated]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "sections_export"

    def get(self, request):
        table, err = _safe_table(request, source="query")
        if err:
            return err
        scope = request.query_params.get("scope", "user")
        handler = EXPORT_SCOPE_REGISTRY.get(scope)
        if handler is None:
            return Response(
                {
                    "detail": f"Unknown export scope {scope!r}.",
                    "allowed_scopes": get_allowed_export_scopes(),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        UserModel = TABLE_TO_USER_MODEL[table]
        count = UserModel.objects.filter(user=request.user, is_active=True).count()
        if count > SECTION_EXPORT_MAX_ROWS:
            return Response(
                {
                    "detail": (
                        f"Too many rows to export ({count}); "
                        f"maximum is {SECTION_EXPORT_MAX_ROWS}."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return handler(table, request.user)
